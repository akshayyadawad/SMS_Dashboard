"""
SMS Campaign — Consolidated Excel Report Generator
Cross-references SmartCard/Renewal SMS data with ACTIVE_Registered_Labours_Data.csv
to produce detailed reports broken down by Labour Officer (LO).

Generates:
  1. Summary Dashboard
  2. SmartCard by LO (day-wise detail)
  3. Renewal by LO (day-wise detail)
  4. Raw data sheets
  5. dashboard_data.json for HTML dashboard
"""

import os
import csv
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import OrderedDict, defaultdict
from datetime import datetime
from functools import lru_cache

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SC_DIR = os.path.join(BASE_DIR, "SmartCard")
SC_MASTER_FILE = os.path.join(BASE_DIR, "New_Smartcard_Data.xlsx")
RENEWAL_FILE = os.path.join(BASE_DIR, "Renewal_30days_from_Apr18.xlsx")
ACTIVE_FILE = os.path.join(BASE_DIR, "ACTIVE_Registered_Labours_Data.csv")
OUTPUT_FILE = os.path.join(BASE_DIR, "SMS_Consolidated_Report.xlsx")
JSON_FILE = os.path.join(BASE_DIR, "dashboard_data.json")

# --- Styling constants ---
DARK_BLUE = "1B2A4A"
MED_BLUE = "2D4373"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=12)
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
sub_header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
sub_header_fill = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
title_font = Font(name="Calibri", bold=True, size=16, color=DARK_BLUE)
subtitle_font = Font(name="Calibri", bold=True, size=13, color=MED_BLUE)
bold_font = Font(name="Calibri", bold=True, size=11)
normal_font = Font(name="Calibri", size=11)
number_font = Font(name="Calibri", size=11)
total_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
total_font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
center = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
right_align = Alignment(horizontal="right", vertical="center")

# SmartCard file → LO file label mapping (for reference/fallback)
LO_NAME_MAP = OrderedDict([
    ("LO1_Makali", "LO1 – Makali"),
    ("LO2_Kengeri", "LO2 – Kengeri"),
    ("LO3_Marathalli", "LO3 – Marathalli"),
    ("LO4_HBR_Layout", "LO4 – HBR Layout"),
    ("LO5_RT_Nagar", "LO5 – RT Nagar"),
    ("LO6_Chandapura", "LO6 – Chandapura"),
    ("LO7_Shanti Nagar", "LO7 – Shanti Nagar"),
    ("LO7_Shanti_Nagar", "LO7 – Shanti Nagar"),
    ("LO8_Jakkur", "LO8 – Jakkur"),
    ("LO_Bng_Rural_Davanahalli", "LO – Devanahalli (Bng Rural)"),
    ("LO_Bng_Rural_Devanahalli", "LO – Devanahalli (Bng Rural)"),
    ("LO_Mysore", "LO – Mysore"),
    ("Mysore", "LO – Mysore"),
])

DAY_FOLDERS = ["28th April", "29th April", "30th April", "3rd May", "4th May", "5th May", "6th May"]


def style_header_row(ws, row, max_col, font=header_font, fill=header_fill):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = center
        cell.border = thin_border


def auto_width(ws, min_width=12, max_width=40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                lengths.append(len(str(cell.value)))
        if lengths:
            best = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = best


# --- LO Name Formatting ---
@lru_cache(maxsize=128)
def format_lo_name(raw_name):
    if not raw_name or raw_name == "Unmatched":
        return raw_name
    
    bng_mapping = {
        "LO1@Bng": "Makali",
        "LO2@Bng": "Kengeri",
        "LO3@Bng": "Marathalli",
        "LO4@Bng": "HBR Layout",
        "LO5@Bng": "RT Nagar",
        "LO6@Bng": "Chandapura",
        "LO7@Bng": "Shanti Nagar",
        "LO8@Bng": "Jakkur"
    }
    
    short = raw_name.split('@')[0]
    
    if raw_name in bng_mapping:
        return f"{raw_name} ({short} {bng_mapping[raw_name]})"
    
    # For others like LO@Mysore -> LO@Mysore (LO Mysore)
    display = raw_name.replace("@", " ")
    if display.startswith("LO") and len(display) > 2 and display[2].isdigit():
        return f"{raw_name} ({short} {raw_name.split('@')[-1]})"
        
    return f"{raw_name} ({display})"



# ─── Data Loading ───

def load_sc_master_data():
    """Load New_Smartcard_Data.xlsx and build mobile to LO lookup."""
    mobile_to_lo = {}
    wb = openpyxl.load_workbook(SC_MASTER_FILE, read_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        lo = str(row[5]).strip() if row[5] else ""
        mob = str(row[8]).strip() if row[8] else ""
        if mob and lo:
            mobile_to_lo[mob] = lo
    wb.close()
    return mobile_to_lo

def load_active_data():
    """Load ACTIVE_Registered_Labours_Data.csv and build lookup for Renewal."""
    mobile_to_lo = {}

    if not os.path.exists(ACTIVE_FILE):
        return {}

    with open(ACTIVE_FILE, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        try:
            headers = next(reader)
        except StopIteration:
            return {}

        mob_idx = -1
        lo_idx = -1
        for idx, h in enumerate(headers):
            h_clean = h.strip().lower()
            if h_clean in ("mobile_no", "mobile no"):
                mob_idx = idx
            elif h_clean in ("labour_officer", "labour officer"):
                lo_idx = idx

        if mob_idx == -1 or lo_idx == -1:
            return {}

        for row in reader:
            if len(row) > max(mob_idx, lo_idx):
                mob = row[mob_idx].strip()
                lo = row[lo_idx].strip()
                if mob and lo:
                    mobile_to_lo[mob] = lo
    return mobile_to_lo


def read_smartcard_data(mobile_to_lo):
    """Read SmartCard files and aggregate by Labour Officer."""
    # lo_day_counts: {lo: {day: count}}
    lo_day_counts = defaultdict(lambda: defaultdict(int))
    # file-level counts for reference
    file_day_counts = defaultdict(lambda: defaultdict(int))
    unmatched_day = defaultdict(int)
    total_per_day = defaultdict(int)

    for day in DAY_FOLDERS:
        day_path = os.path.join(SC_DIR, day)
        if not os.path.isdir(day_path):
            continue
        for fname in sorted(os.listdir(day_path)):
            if not fname.endswith(".xlsx") or fname.startswith("."):
                continue
            key = fname.replace(".xlsx", "")
            file_label = LO_NAME_MAP.get(key, key)

            fpath = os.path.join(day_path, fname)
            wb = openpyxl.load_workbook(fpath, read_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                mob_raw = row[0]
                mob = str(int(mob_raw)) if isinstance(mob_raw, (int, float)) else str(mob_raw).strip()
                lo_raw = mobile_to_lo.get(mob, "")
                total_per_day[day] += 1
                if lo_raw:
                    lo = format_lo_name(lo_raw)
                    lo_day_counts[lo][day] += 1
                else:
                    unmatched_day[day] += 1
                file_day_counts[file_label][day] += 1
            wb.close()

    return dict(lo_day_counts), dict(file_day_counts), dict(unmatched_day), dict(total_per_day)


def read_renewal_data(mobile_to_lo):
    """Read Renewal data and aggregate by Labour Officer and date."""
    wb = openpyxl.load_workbook(RENEWAL_FILE, read_only=True)
    ws = wb.active
    headers = None

    # lo_date_counts: {lo: {date_str: count}}
    lo_date_counts = defaultdict(lambda: defaultdict(int))
    unmatched = 0
    total = 0
    all_dates = set()

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        total += 1
        mob = str(row[2]).strip() if row[2] else ""
        created = row[8]
        if hasattr(created, 'date'):
            date_str = str(created.date())
        elif isinstance(created, str):
            date_str = created.split(' ')[0]
        else:
            date_str = "Unknown"
        all_dates.add(date_str)

        lo_raw = mobile_to_lo.get(mob, "")
        if lo_raw:
            lo = format_lo_name(lo_raw)
            lo_date_counts[lo][date_str] += 1
        else:
            lo_date_counts["Unmatched"][date_str] += 1
            unmatched += 1

    wb.close()
    sorted_dates = sorted(all_dates)
    return dict(lo_date_counts), sorted_dates, total, unmatched


# ─── Excel Sheet Builders ───

def create_summary_sheet(wb, sc_lo_data, sc_file_data, sc_unmatched, sc_totals,
                         rn_lo_data, rn_dates, rn_total, rn_unmatched):
    """Create the Summary Dashboard sheet."""
    ws = wb.active
    ws.title = "Summary Dashboard"
    ws.sheet_properties.tabColor = DARK_BLUE

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "SMS Campaign — Consolidated Report"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color=DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Report Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
    ws["A2"].font = Font(name="Calibri", size=11, color="888888")
    ws["A2"].alignment = center

    # Overall Summary
    row = 4
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "OVERALL SUMMARY"
    ws[f"A{row}"].font = subtitle_font
    ws.row_dimensions[row].height = 28

    row = 5
    headers_list = ["Campaign", "Total SMS", "Days", "Labour Officers", "Date Range", "Match Rate", "Status"]
    for c, h in enumerate(headers_list, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers_list))

    sc_grand = sum(sc_totals.values())
    sc_matched = sc_grand - sum(sc_unmatched.values())
    sc_match_pct = f"{sc_matched/sc_grand*100:.1f}%" if sc_grand else "0%"
    sc_los = len(sc_lo_data)

    rn_matched = rn_total - rn_unmatched
    rn_match_pct = f"{rn_matched/rn_total*100:.1f}%" if rn_total else "0%"
    rn_los = len([k for k in rn_lo_data if k != "Unmatched"])

    row = 6
    for c, v in enumerate(["SmartCard", sc_grand, len(DAY_FOLDERS), f"{sc_los} LOs",
                            "28 Apr – 3 May 2026", sc_match_pct, "Completed"], 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = normal_font; cell.alignment = center; cell.border = thin_border
        if c == 2: cell.number_format = '#,##0'

    row = 7
    for c, v in enumerate(["Renewal 30 Days", rn_total, len(rn_dates), f"{rn_los} LOs",
                            f"{rn_dates[0]} to {rn_dates[-1]}", rn_match_pct, "Completed"], 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = normal_font; cell.alignment = center; cell.border = thin_border
        if c == 2: cell.number_format = '#,##0'
        cell.fill = alt_fill

    row = 8
    for c, v in enumerate(["GRAND TOTAL", sc_grand + rn_total, "—", "—", "—", "—", "All Done"], 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = total_font; cell.fill = total_fill; cell.alignment = center; cell.border = thin_border
        if c == 2: cell.number_format = '#,##0'

    # SmartCard by LO
    row = 10
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "SMARTCARD SMS — BY LABOUR OFFICER"
    ws[f"A{row}"].font = subtitle_font
    ws.row_dimensions[row].height = 28

    row = 11
    sc_headers = ["#", "Labour Officer"] + DAY_FOLDERS + ["Total"]
    for c, h in enumerate(sc_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(sc_headers))

    row = 12
    day_totals = defaultdict(int)
    grand = 0
    sorted_los = sorted(sc_lo_data.items(), key=lambda x: -sum(x[1].values()))
    for idx, (lo, day_counts) in enumerate(sorted_los, 1):
        ws.cell(row=row, column=1, value=idx).font = normal_font
        ws.cell(row=row, column=2, value=lo).font = bold_font
        lo_total = 0
        for c, day in enumerate(DAY_FOLDERS, 3):
            cnt = day_counts.get(day, 0)
            cell = ws.cell(row=row, column=c, value=cnt)
            cell.number_format = '#,##0'; cell.font = number_font; cell.alignment = right_align
            day_totals[day] += cnt; lo_total += cnt
        cell = ws.cell(row=row, column=len(sc_headers), value=lo_total)
        cell.number_format = '#,##0'; cell.font = total_font; cell.alignment = right_align
        grand += lo_total
        for c in range(1, len(sc_headers)+1):
            ws.cell(row=row, column=c).border = thin_border
            if idx % 2 == 0: ws.cell(row=row, column=c).fill = alt_fill
        row += 1

    # Unmatched row
    um_total = sum(sc_unmatched.values())
    if um_total > 0:
        ws.cell(row=row, column=1, value="").font = normal_font
        ws.cell(row=row, column=2, value="Unmatched").font = Font(name="Calibri", bold=True, color="E74C3C")
        for c, day in enumerate(DAY_FOLDERS, 3):
            cell = ws.cell(row=row, column=c, value=sc_unmatched.get(day, 0))
            cell.number_format = '#,##0'; cell.font = number_font; cell.alignment = right_align
        cell = ws.cell(row=row, column=len(sc_headers), value=um_total)
        cell.number_format = '#,##0'; cell.font = total_font
        for c in range(1, len(sc_headers)+1):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    # Grand total
    ws.cell(row=row, column=2, value="GRAND TOTAL").font = total_font
    for c, day in enumerate(DAY_FOLDERS, 3):
        cell = ws.cell(row=row, column=c, value=day_totals[day] + sc_unmatched.get(day, 0))
        cell.number_format = '#,##0'; cell.font = total_font; cell.alignment = right_align
    cell = ws.cell(row=row, column=len(sc_headers), value=grand + um_total)
    cell.number_format = '#,##0'; cell.font = total_font; cell.alignment = right_align
    for c in range(1, len(sc_headers)+1):
        ws.cell(row=row, column=c).fill = total_fill; ws.cell(row=row, column=c).border = thin_border

    # Renewal by LO
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "RENEWAL 30 DAYS SMS — BY LABOUR OFFICER"
    ws[f"A{row}"].font = subtitle_font
    ws.row_dimensions[row].height = 28

    row += 1
    rn_headers = ["#", "Labour Officer"] + [d.replace("2026-", "") for d in rn_dates] + ["Total"]
    for c, h in enumerate(rn_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(rn_headers))

    row += 1
    rn_sorted = sorted(
        [(lo, dc) for lo, dc in rn_lo_data.items() if lo != "Unmatched"],
        key=lambda x: -sum(x[1].values())
    )
    rn_day_totals = defaultdict(int)
    rn_grand = 0
    for idx, (lo, date_counts) in enumerate(rn_sorted, 1):
        ws.cell(row=row, column=1, value=idx).font = normal_font
        ws.cell(row=row, column=2, value=lo).font = bold_font
        lo_total = 0
        for c, dt in enumerate(rn_dates, 3):
            cnt = date_counts.get(dt, 0)
            cell = ws.cell(row=row, column=c, value=cnt)
            cell.number_format = '#,##0'; cell.font = number_font; cell.alignment = right_align
            rn_day_totals[dt] += cnt; lo_total += cnt
        cell = ws.cell(row=row, column=len(rn_headers), value=lo_total)
        cell.number_format = '#,##0'; cell.font = total_font; cell.alignment = right_align
        rn_grand += lo_total
        for c in range(1, len(rn_headers)+1):
            ws.cell(row=row, column=c).border = thin_border
            if idx % 2 == 0: ws.cell(row=row, column=c).fill = alt_fill
        row += 1

    # Grand total
    ws.cell(row=row, column=2, value="GRAND TOTAL").font = total_font
    for c, dt in enumerate(rn_dates, 3):
        cell = ws.cell(row=row, column=c, value=rn_day_totals[dt])
        cell.number_format = '#,##0'; cell.font = total_font; cell.alignment = right_align
    cell = ws.cell(row=row, column=len(rn_headers), value=rn_grand)
    cell.number_format = '#,##0'; cell.font = total_font; cell.alignment = right_align
    for c in range(1, len(rn_headers)+1):
        ws.cell(row=row, column=c).fill = total_fill; ws.cell(row=row, column=c).border = thin_border

    auto_width(ws)
    ws.freeze_panes = "A6"


def generate_dashboard_json(sc_lo_data, sc_file_data, sc_unmatched, sc_totals,
                            rn_lo_data, rn_dates, rn_total, rn_unmatched):
    """Generate JSON data file for the HTML dashboard."""
    # Map full folder names to short keys for dashboard
    folder_map = {
        "28th April": "2026-04-28",
        "29th April": "2026-04-29",
        "30th April": "2026-04-30",
        "3rd May": "2026-05-03",
        "4th May": "2026-05-04",
        "5th May": "2026-05-05",
        "6th May": "2026-05-06"
    }

    # SmartCard by LO
    sc_rows = []
    for lo, day_counts in sorted(sc_lo_data.items(), key=lambda x: -sum(x[1].values())):
        row = {"lo": lo}
        total = 0
        for day in DAY_FOLDERS:
            cnt = day_counts.get(day, 0)
            key = folder_map.get(day, day)
            row[key] = cnt
            total += cnt
        row["total"] = total
        sc_rows.append(row)

    # Add unmatched
    um_total = sum(sc_unmatched.values())
    if um_total > 0:
        um_row = {"lo": "Unmatched"}
        for day in DAY_FOLDERS:
            key = folder_map.get(day, day)
            um_row[key] = sc_unmatched.get(day, 0)
        um_row["total"] = um_total
        sc_rows.append(um_row)

    sc_grand = sum(sc_totals.values())

    # SmartCard by file (original grouping)
    sc_file_rows = []
    for label, day_counts in sorted(sc_file_data.items(), key=lambda x: -sum(x[1].values())):
        row = {"loc": label}
        total = 0
        for day in DAY_FOLDERS:
            cnt = day_counts.get(day, 0)
            key = folder_map.get(day, day)
            row[key] = cnt
            total += cnt
        row["total"] = total
        sc_file_rows.append(row)

    # Renewal by LO
    rn_rows = []
    for lo, date_counts in sorted(
        rn_lo_data.items(),
        key=lambda x: -sum(x[1].values())
    ):
        row = {"lo": lo}
        total = 0
        for dt in rn_dates:
            cnt = date_counts.get(dt, 0)
            row[dt] = cnt
            total += cnt
        row["total"] = total
        rn_rows.append(row)

    data = {
        "generated": datetime.now().strftime("%d %B %Y"),
        "smartcard": {
            "days": DAY_FOLDERS,
            "by_lo": sc_rows,
            "by_file": sc_file_rows,
            "grand_total": sc_grand,
            "lo_count": len(sc_lo_data),
        },
        "renewal": {
            "dates": rn_dates,
            "by_lo": rn_rows,
            "grand_total": rn_total,
            "lo_count": len([k for k in rn_lo_data if k != "Unmatched"]),
            "days_count": len(rn_dates),
        }
    }

    with open(JSON_FILE, "w") as f:
        json.dump(data, f, indent=2)

    return data


def main():
    print("=" * 60)
    print("  SMS Campaign — Consolidated Report Generator")
    print("  Cross-referencing with ACTIVE Registered Labourers")
    print("=" * 60)

    # 1. Load Master Data
    print("1. Loading Master Data...")
    sc_master_mapping = load_sc_master_data()
    print(f"   SmartCard master: {len(sc_master_mapping)} entries")
    rn_mobile_to_lo = load_active_data()
    print(f"   Renewal master: {len(rn_mobile_to_lo)} entries")

    # 2. Reading SmartCard SMS data
    print("\n2. Reading SmartCard SMS data...")
    sc_lo_data, sc_file_data, sc_unmatched, sc_totals = read_smartcard_data(sc_master_mapping)
    sc_grand = sum(sc_totals.values())
    sc_matched = sc_grand - sum(sc_unmatched.values())
    print(f"   Total SMS: {sc_grand:,}")
    print(f"   Matched to LO: {sc_matched:,} ({sc_matched/sc_grand*100:.1f}%)")
    print(f"   Unmatched: {sum(sc_unmatched.values()):,}")
    print(f"   Unique LOs found: {len(sc_lo_data)}")

    # 3. Reading Renewal data
    print("\n3. Reading Renewal 30 Days data...")
    rn_lo_data, rn_dates, rn_total, rn_unmatched = read_renewal_data(rn_mobile_to_lo)
    rn_matched = rn_total - rn_unmatched
    print(f"   Total SMS: {rn_total:,}")
    print(f"   Matched to LO: {rn_matched:,} ({rn_matched/rn_total*100:.1f}%)")
    print(f"   Unmatched: {rn_unmatched:,}")
    print(f"   Unique LOs: {len([k for k in rn_lo_data if k != 'Unmatched'])}")
    print(f"   Date range: {rn_dates[0]} to {rn_dates[-1]} ({len(rn_dates)} days)")

    # Create Excel
    print("\n4. Creating Excel report...")
    wbk = openpyxl.Workbook()
    create_summary_sheet(wbk, sc_lo_data, sc_file_data, sc_unmatched, sc_totals,
                         rn_lo_data, rn_dates, rn_total, rn_unmatched)
    print(f"   Saving to: {OUTPUT_FILE}")
    wbk.save(OUTPUT_FILE)
    print(f"   Size: {os.path.getsize(OUTPUT_FILE) / (1024*1024):.1f} MB")

    # Generate JSON for dashboard
    print("\n5. Generating dashboard JSON...")
    data = generate_dashboard_json(sc_lo_data, sc_file_data, sc_unmatched, sc_totals,
                                   rn_lo_data, rn_dates, rn_total, rn_unmatched)
    print(f"   Saved to: {JSON_FILE}")
    print(f"   SmartCard LOs: {data['smartcard']['lo_count']}")
    print(f"   Renewal LOs: {data['renewal']['lo_count']}")

    print("\n" + "=" * 60)
    print("  Report generated successfully!")
    print("=" * 60)


if __name__ == "__main__":
    main()